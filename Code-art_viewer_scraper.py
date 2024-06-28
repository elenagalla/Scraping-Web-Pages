from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
import pandas as pd
import re
import time
import requests
from tqdm import tqdm
import concurrent.futures
import logging
import urllib.parse
import json
import gc  # for garbage collection
from openpyxl import load_workbook

# Define the URLs
urls = [
    "https://artviewer.org/2024/05/",
    # Add more URLs as needed
]

# List to hold all exhibition data from all URLs
all_data = []
batch_size = 10  # Define batch size

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Helper function to get text after a label
def get_text_after_label(label):
    if label:
        sibling = label.next_sibling
        while sibling and (not isinstance(sibling, str) or sibling.strip() == ''):
            sibling = sibling.next_sibling
        if sibling and isinstance(sibling, str):
            return sibling.strip()
    return 'N/A'

# Function to scrape additional data from nested URLs
def scrape_nested_url(url, extracted_info):
    logging.info(f"Scraping nested URL: {url}")
    try:
        for attempt in range(3):  # Retry mechanism
            try:
                response = requests.get(url)
                response.raise_for_status()
                break
            except requests.RequestException as e:
                logging.warning(f"Attempt {attempt + 1} failed for {url}. Error: {e}")
                time.sleep(2)
        else:
            logging.error(f"Failed to retrieve the page {url} after multiple attempts.")
            return 'N/A', 'N/A', 'N/A', 'N/A', []

        soup = BeautifulSoup(response.text, 'html.parser')
        entry_content_div = soup.find('div', class_='entry-content')
        if not entry_content_div:
            logging.info(f"No entry content found for {url}")
            return 'N/A', 'N/A', 'N/A', 'N/A', []

        paragraphs = entry_content_div.find_all('p')
        filtered_paragraphs = [p.get_text(strip=True) for p in paragraphs if not any(
            text in p.get_text(strip=True) for text in ['Artist:', 'Artists:', 'Exhibition title:', 'Venue:', 'Date:', 'Photography:'])]
        paragraph_text = '\n'.join(filtered_paragraphs)
        images = [img['src'] for img in entry_content_div.find_all('img')]

        date, venue, exhibition_title = 'N/A', 'N/A', 'N/A'
        for p in paragraphs:
            text = p.get_text(strip=True)
            if 'Date:' in text:
                date = text.split('Date:')[-1].strip()
            if 'Venue:' in text:
                venue = text.split('Venue:')[-1].strip()
            if 'Exhibition Title:' in text:
                exhibition_title = text.split('Exhibition Title:')[-1].strip()

        if paragraph_text and paragraph_text not in extracted_info:
            extracted_info.append(paragraph_text)
            return paragraph_text, date, venue, exhibition_title, images
        return 'N/A', date, venue, exhibition_title, images
    except Exception as e:
        logging.error(f"Failed to scrape nested URL {url}. Error: {e}")
        return 'N/A', 'N/A', 'N/A', 'N/A', []

def load_lazy_content(driver):
    SCROLL_PAUSE_TIME = 0.1  # Reduced pause time to make scrolling faster
    SCROLL_INCREMENT = 300
    MAX_ATTEMPTS = 10
    no_change_attempts = 0

    last_height = driver.execute_script("return document.body.scrollHeight")

    while no_change_attempts < MAX_ATTEMPTS:
        driver.execute_script(f"window.scrollBy(0, {SCROLL_INCREMENT});")
        time.sleep(SCROLL_PAUSE_TIME)
        new_height = driver.execute_script("return document.body.scrollHeight")
        logging.info(f"Scrolled to: {new_height}")
        if new_height == last_height:
            no_change_attempts += 1
        else:
            no_change_attempts = 0
        last_height = new_height

    time.sleep(1)  # Ensure all content is fully loaded

def scrape_url(url):
    options = webdriver.ChromeOptions()
    options.add_argument("--headless")
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    try:
        logging.info(f"Scraping URL: {url}")
        driver.get(url)
        load_lazy_content(driver)
        WebDriverWait(driver, 20).until(EC.presence_of_all_elements_located((By.CLASS_NAME, "post")))

        soup = BeautifulSoup(driver.page_source, 'html.parser')
        title = soup.title.string if soup.title else 'No title'
        logging.info(f"Retrieved title from {url}: {title}")

        exhibitions = soup.find_all('article', class_='post')
        logging.info(f"Found {len(exhibitions)} exhibitions on the page.")
        data = []

        for exhibition in exhibitions:
            artists, img_urls, additional_images = [], [], []
            additional_info = ''
            extracted_info = []  # Initialize a new list for each exhibition

            # Initialize exhibition data with N/A
            exhibition_title = 'N/A'
            venue = 'N/A'
            date = 'N/A'

            strong_tags = exhibition.find_all('strong')
            for tag in strong_tags:
                tag_text = tag.get_text(strip=True)
                if 'Exhibition title:' in tag_text:
                    exhibition_title = get_text_after_label(tag)
                elif 'Artist:' in tag_text or 'Artists:' in tag_text:
                    potential_artists = get_text_after_label(tag)
                    if potential_artists.lower() != 'n/a' and not any(char.isdigit() for char in potential_artists):
                        if not re.search(r'https?://', potential_artists):
                            artists.extend([artist.strip() for artist in re.split(r'[,|&]', potential_artists) if artist.strip()])
                elif 'Date:' in tag_text:
                    date = get_text_after_label(tag)

            title_tag = exhibition.find('h2', class_='entry-title')
            if title_tag:
                title_text = title_tag.get_text(strip=True)
                if ' at ' in title_text:
                    venue = title_text.split(' at ')[-1].strip()
                    if not artists:
                        artist_match = re.search(r'^(.*?) at ', title_text)
                        if artist_match:
                            artists_text = artist_match.group(1)
                            artists.extend([artist.strip() for artist in re.split(r'[,|&]', artists_text) if artist.strip()])

            img_tag = exhibition.find('img')
            if (img_tag and 'src' in img_tag.attrs):
                img_urls.append(img_tag['src'])

            # Scrape nested URLs for additional info and images
            nested_links = [a['href'] for a in exhibition.find_all('a', href=True) if a['href'].startswith("https://artviewer.org/")]
            for nested_url in nested_links:
                paragraph, nested_date, nested_venue, nested_exhibition_title, nested_images = scrape_nested_url(nested_url, extracted_info)
                if paragraph != 'N/A':
                    additional_info += paragraph + '\n'
                additional_images.extend(nested_images)
                if date == 'N/A' and nested_date != 'N/A':
                    date = nested_date
                if venue == 'N/A' and nested_venue != 'N/A':
                    venue = nested_venue
                if exhibition_title == 'N/A' and nested_exhibition_title != 'N/A':
                    exhibition_title = nested_exhibition_title

            img_urls.extend(additional_images)
            img_urls = list(set(img_urls))

            # Save URLs as a JSON string
            img_urls_string = json.dumps(img_urls)

            # Print length of img_urls_string for debugging
            logging.info(f"Length of img_urls_string: {len(img_urls_string)}")

            # Log additional info length
            logging.info(f"Length of additional_info: {len(additional_info)}")

            # Ensure the additional info is properly encoded
            additional_info = additional_info.encode('utf-8', errors='ignore').decode('utf-8', errors='ignore')

            exhibition_data = {
                'Artist': ', '.join(artists) if artists else 'N/A',
                'Venue': venue,
                'Exhibition title': exhibition_title,
                'Date': date,
                'Image_URLs': img_urls_string,
                'Additional Info': additional_info.strip()
            }

            logging.info(f"Exhibition data: {exhibition_data}")
            data.append(exhibition_data)

        return data
    except Exception as e:
        logging.error(f"Failed to retrieve the page {url}. Error: {e}")
        return []
    finally:
        driver.quit()
        gc.collect()  # Explicitly call garbage collection to free up memory

def save_data(df, filename, mode='a', header=True):
    try:
        logging.info("DataFrame content before saving:")
        logging.info(df.to_string())
        
        with pd.ExcelWriter(filename, engine='openpyxl', mode=mode) as writer:
            if 'Sheet1' in writer.book.sheetnames:
                startrow = writer.book['Sheet1'].max_row
            else:
                startrow = 0
            df.to_excel(writer, index=False, header=header, startrow=startrow)
        
        logging.info(f"Data has been saved to {filename}")
    except Exception as e:
        logging.error(f"Failed to save data to {filename}. Error: {e}")

def process_batches(data, batch_size, filename):
    for i in range(0, len(data), batch_size):
        batch = data[i:i + batch_size]
        df = pd.DataFrame(batch)
        header = True if i == 0 else False  # Write header only for the first batch
        mode = 'w' if i == 0 else 'a'  # Write mode for the first batch, append mode for subsequent batches
        save_data(df, filename, mode=mode, header=header)

def main():
    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
        results = list(tqdm(executor.map(scrape_url, urls), total=len(urls), desc="Scraping URLs"))
    
    for result in results:
        if isinstance(result, list):
            all_data.extend(result)
        else:
            logging.error(f"Error occurred: {result}")

    # Process data in batches
    process_batches(all_data, batch_size, 'data.xlsx')

if __name__ == "__main__":
    main()
